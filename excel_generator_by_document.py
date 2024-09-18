# 功能excel生成脚本，适用于有功能概述的项目

import re
from metagpt.actions import Action
import asyncio
import os
import pandas as pd
from metagpt.roles import Role
from metagpt.context import Context
from metagpt.logs import logger
from metagpt.schema import Message
import typer
from metagpt.team import Team
from openpyxl import load_workbook

app = typer.Typer()

os.environ["http_proxy"] = "http://localhost:7890"
os.environ["https_proxy"] = "http://localhost:7890"

# Read the product description from a file
def read_description_from_file(file_path: str):
    with open(file_path, 'r', encoding='utf-8') as file:
        description = file.read().strip()
    return description

# Generate modules based on the product description
class GenerateModules(Action):
    PROMPT_TEMPLATE: str = """
    Based on the product description: '{description}', generate a list of modules based on the description.
    No other except modules should be output, and output should be in Chinese.
    Example:
    '1. **数据摄取模块：**', 
    2. **数据存储与管理模块：**'
    3. **数据处理与分析模块：**'
    """
    name: str = "GenerateModules"

    async def run(self, description: str):
        prompt = self.PROMPT_TEMPLATE.format(description=description)
        rsp = await self._aask(prompt)
        logger.info(f"生成模块: {rsp}")
        modules = [module.strip() for module in rsp.split("\n") if module.strip()]
        return modules

class GenerateModuleDetails(Action):
    PROMPT_TEMPLATE: str = (
      """For the module '{module}', provide a detailed list of functions based on document: '{description}'.
      u should see that the module is likely one of the item in the document, and following is some discription which is helpful for generating functions
      For each function, include function pages and a brief function overview.
      Requirements: 
      Output should be in Chinese except for keywords like Module, Function, Function Page, Function Overview.
      Each module can have multiple functions.
      Each function has a name like Data source management.
      Each function has at least 4  function pages, such as Add data source page.
      Each function page has a brief overview like adding a data source with a graphical interface for users to input database details (IP, port, etc.).
      Example:
      
      Module：数据集成和预处理

      Function：数据源管理
        Function Page：添加数据源页面
        Function Overview：添加数据源，并提供图形界面供用户输入数据源的基本信息（如数据库类型、IP地址、端口号、用户名、密码）和连接参数。
        
        Function Page：数据源列表页面
        Function Overview：显示已添加的数据源列表，允许用户编辑、删除或测试每个数据源的连接。
          
      Function：数据清洗
        Function Page：清洗规则配置页面
        Function Overview：提供接口供用户定义各种清理规则（如删除空值、替换错误值、设置数据类型约束）。
      """
    )
    name: str = "GenerateModuleDetails"

    async def run(self, module: str, description: str):
        prompt = self.PROMPT_TEMPLATE.format(module=module, description=description)
        rsp = await self._aask(prompt)
        logger.info(f"Generated details for module {module}: {rsp}")
        details = [detail.strip() for detail in rsp.split("\n") if detail.strip()]
        return details


# Save the generated requirements to an Excel file
class SaveRequirementsToExcel(Action):
    name: str = "SaveRequirementsToExcel"

    async def run(self, requirements_data: list, filename="需求模块.xlsx"):
        self.save_to_excel(requirements_data, filename)
        logger.info(f"Requirements saved to {filename}")
        await self.clean_and_translate_excel(filename)

    def save_to_excel(self, data, filename="需求模块.xlsx"):
        df = pd.DataFrame(data, columns=["模块", "功能", "功能页面", "功能概述"])
        df.index += 1  # Start the index from 1 for better readability in Excel
        df.index.name = 'ID'
        
        # Define default column widths (no 'ID' column in the DataFrame)
        column_widths = {
            "模块": 30,
            "功能": 30,
            "功能页面": 40,
            "功能概述": 100
        }
        
        if os.path.exists(filename):
            # Append to the existing Excel file
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                start_row = writer.sheets['Sheet1'].max_row  # Get the current max row
                df.to_excel(writer, startrow=start_row, header=False, index=False)

                # Access the workbook and worksheet to adjust column widths
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                for col_name, width in column_widths.items():
                    col_letter = chr(ord('A') + df.columns.get_loc(col_name))
                    worksheet.column_dimensions[col_letter].width = width
        else:
            # Create a new Excel file
            df.to_excel(filename, engine='openpyxl', index=False)
            
            # Load the workbook to modify column widths
            workbook = load_workbook(filename)
            worksheet = workbook.active
            for col_name, width in column_widths.items():
                col_letter = chr(ord('A') + df.columns.get_loc(col_name))
                worksheet.column_dimensions[col_letter].width = width
            
            workbook.save(filename)

    async def clean_and_translate_excel(self, filename):
        # Load the Excel file
        df = pd.read_excel(filename)

        # Remove incomplete rows
        df_cleaned = df.dropna(subset=["模块", "功能", "功能页面", "功能概述"])

        # Save the cleaned and translated data back to the Excel file
        df_cleaned.to_excel(filename, index=False)
        logger.info(f"数据已清洗翻译并保存于 {filename}")

# The main role for the requirements engineering process
class AutomatedRequirementsEngineer(Role):
    name: str = "AutomatedRequirementsEngineer"

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.set_actions([GenerateModules, GenerateModuleDetails, SaveRequirementsToExcel])

    async def _act(self) -> Message:
        # Read product description from file
        product_description_file = "description.txt"  # You can change the file name or path here
        product_description = read_description_from_file(product_description_file)
        logger.info(f"Product description loaded....")


        # Step 1: Generate Modules
        generate_modules_action = GenerateModules()
        generated_response = await generate_modules_action.run(description=product_description)
    
        # 只提取module名
        modules = [line for line in generated_response if re.match(r"^\d+", line)]
        logger.info(f"生成模块: {modules}")
        logger.info(f"模块生成器正在休眠")
        await asyncio.sleep(10)
        logger.info(f"停止休眠")

        requirements_data = []

        for module in modules:
            # Step 2: Generate details for each module
            generate_module_details_action = GenerateModuleDetails()
            module_details = await generate_module_details_action.run(module=module, description=product_description)

            # Parse and structure the response into requirements data
            parsed_details = self.parse_module_details("\n".join(module_details))
            requirements_data.extend(parsed_details)

            logger.info(f"正在保存模块: {module}")
            save_action = SaveRequirementsToExcel()
            logger.info(f"保存成功")
            await save_action.run(requirements_data, filename="企业信息及流程智能化服务平台.xlsx")
            requirements_data.clear()
            logger.info(f"功能生成器正在休眠")
            await asyncio.sleep(30)
            logger.info(f"停止休眠")

        msg = Message(content="Requirements generated and saved to Excel.", role=self.profile, cause_by=type(generate_modules_action))
        return msg

    def parse_module_details(self, detail: str):
        module_data = []
        current_module = None
        current_function = None
        current_function_page = None

        lines = detail.split("\n")
        for line in lines:
          line = line.strip()

          if line.startswith("Module："):
              current_module = line.replace("Module：", "").strip()
              
          elif line.startswith("Function："):
              current_function = line.replace("Function：", "").strip()
          elif line.startswith("Function Page："):
              current_function_page = line.replace("Function Page：", "").strip()
          elif line.startswith("Function Overview："):
              function_overview = line.replace("Function Overview：", "").strip()
              # Add the collected details to the list
              module_data.append({
                  "模块": current_module,
                  "功能": current_function,
                  "功能页面": current_function_page,
                  "功能概述": function_overview,
              })
          else:
            continue
                
        return module_data

@app.command()
def main(
    description_file: str = typer.Argument(..., help="Path to the product description file."),
    investment: float = typer.Option(default=3.0, help="Dollar amount to invest in the AI company."),
    n_round: int = typer.Option(default=5, help="Number of rounds for the simulation."),
):
    logger.info(description_file)

    team = Team()
    team.hire(
        [
            AutomatedRequirementsEngineer(),
        ]
    )

    team.invest(investment=investment)
    logger.info(f"Running project with description from file: {description_file}")

    team.run_project(description_file)
    asyncio.run(team.run(n_round=n_round))

if __name__ == '__main__':
    app()
