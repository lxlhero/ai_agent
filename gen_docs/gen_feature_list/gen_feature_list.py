import pandas as pd
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage
import os
import json
from langchain import hub
from langchain_chroma import Chroma
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain.prompts import PromptTemplate
from langchain_chroma import Chroma
import docx
from typing import Any
import logging
from docx.shared import Pt
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# 根据产品简述生成页面级别的功能清单

# 从环境变量中获取 OPENAI_API_KEY
openai_api_key = os.getenv("OPENAI_API_KEY")

if openai_api_key is None:
    raise ValueError("OPENAI_API_KEY environment variable is not set")
else:
    os.environ["OPENAI_API_KEY"] = str(openai_api_key)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DataPreprocessingModuleSpecGenerator:
    def __init__(self, overview_path):
        # 初始化参数
        self.embeddings = OpenAIEmbeddings()
        self.chroma_db = Chroma(embedding_function=self.embeddings)
        self.overview_path = overview_path

    # 提取并保存采购报价单excel
    def save_func_names_json_to_excel(self, func_names_json, output_file_path):
        """
        将 func_names_json 数据处理成 Excel 文件并保存。

        参数:
        func_names_json (list): 包含功能模块和子功能描述的 JSON 数据。
        output_file_path (str): 输出 Excel 文件的路径。
        """
        # Prepare data for DataFrame
        data = []
        
        modules = func_names_json["modules"]
        for module in modules:
            module_name = module["module_name"]
            for function in module["functions"]:
                function_name = function["function_name"]
                for function_page in function["function_pages"]:
                    function_page_overview = function_page["page_overview"]
                    data.append({"模块": module_name, "功能": function_name, "功能页面": function_page["page_name"], "功能概述": function_page_overview})

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Write DataFrame to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Merge cells with the same "功能模块"
        prev_module_name = None
        start_row = 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            current_module_name = row[0].value
            if current_module_name != prev_module_name:
                if prev_module_name is not None:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=row[0].row-1, end_column=1)
                prev_module_name = current_module_name
                start_row = row[0].row
        if prev_module_name is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)

        # Merge cells with the same "功能"
        prev_function_name = None
        start_row = 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            current_function_name = row[0].value
            if current_function_name != prev_function_name:
                if prev_function_name is not None:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=row[0].row-1, end_column=2)
                prev_function_name = current_function_name
                start_row = row[0].row
        if prev_function_name is not None:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)

        # Set alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths with better handling for wide characters
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get column letter (e.g., 'A', 'B')
            for cell in col:
                if cell.value:
                    # Estimate character width: consider wide characters like Chinese as 2 units
                    value = str(cell.value)
                    adjusted_length = sum(2 if ord(char) > 127 else 1 for char in value)
                    max_length = max(max_length, adjusted_length)
            ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

        # Save the Excel file
        wb.save(output_file_path)

    

    def process(self):
        self.llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.2)

        # 读取产品简述
        with open(self.overview_path, 'r', encoding='utf8') as f:
            product_description = f.read()

        

        # 生成包含功能模块及其子功能描述的JSON
        func_names_json_prompt = (
            "根据以下产品简述，生成页面级的功能清单，包括模块、功能、功能页面和功能概述的JSON。\n"
            "一个产品应有多个模块，每个模块包含多个功能，一个功能有多个功能页面，一个功能页面对应一个功能页面概述。\n"
            "请仔细分析拆分产品简述，每一个模块都要生成，每一个模块的每一行就是一个功能，每一个模块里的功能也必须生成\n"
            "至少生成6个模块, 每个模块至少包含10个及以上的功能，每个功能至少有2个页面，功能页面概述需要详细描述该页面的功能\n"
            "产品简述:\n"
            f"{product_description}\n\n"
            
            "生成的JSON必须严格符合以下格式, 注意这里只有一个模块，只是一个示例，具体的功能要根据产品实际情况定义：\n"
            "{\n"
            "  \"modules\": [\n"
            "    {\n"
            "      \"module_name\": \"平台基础模块\",\n"
            "      \"functions\": [\n"
            "        {\n"
            "          \"function_name\": \"系统菜单管理\",\n"
            "          \"function_pages\": [\n"
            "            {\n"
            "              \"page_name\": \"系统菜单层级管理页面\",\n"
            "              \"page_overview\": \"系统菜单层级管理页面的主要功能为系统菜单配置层级并进行管理，以便管理员或具有相应权限的用户能够轻松地查看和管理菜单结构，确保菜单的层次清晰，便于用户导航。\"\n"
            "            },\n"
            "            {\n"
            "              \"page_name\": \"系统菜单管理页面\",\n"
            "              \"page_overview\": \"系统菜单管理页面的主要功能为录入系统所有的菜单页面，并配置访问路径、页面操作按钮、菜单层级、标签等信息，以确保菜单的完整性和可用性。\"\n"
            "            },\n"
            "            {\n"
            "              \"page_name\": \"系统菜单访问记录管理\",\n"
            "              \"page_overview\": \"系统菜单访问记录管理的主要功能为查询菜单的用户访问记录、访问时间和访问地址等信息，以便进行用户行为分析、安全审计和性能优化。\"\n"
            "            }\n"
            "          ]\n"
            "        },\n"
            "        {\n"
            "          \"function_name\": \"系统元数据管理\",\n"
            "          \"function_pages\": [\n"
            "            {\n"
            "              \"page_name\": \"元数据管理页面\",\n"
            "              \"page_overview\": \"元数据管理页面主要是用来配置系统中交通运输、车辆、气象元数据信息，建立元数据目录和配置数据依赖关系，作为系统数据权限和数据服务的重要依据，确保数据的准确性和一致性。\"\n"
            "            },\n"
            "            {\n"
            "              \"page_name\": \"元数据关联页面\",\n"
            "              \"page_overview\": \"元数据关联页面主要功能是将交通大数据平台中的交通、路网、车辆、气象等元数据进行关联管理，为交通大数据平台后续业务提供基础数据支撑，提高数据分析的深度和广度。\"\n"
            "            },\n"
            "            {\n"
            "              \"page_name\": \"元数据版本管理页面\",\n"
            "              \"page_overview\": \"元数据版本管理页面的主要功能在于对元数据的版本进行有效的管理和控制，确保元数据的变更历史清晰可追溯，同时支持用户根据需要进行版本比对和恢复，保障数据的安全性和完整性。\"\n"
            "            },\n"
            "            {\n"
            "              \"page_name\": \"元数据权限页面\",\n"
            "              \"page_overview\": \"元数据权限页面的主要功能在于对元数据进行权限控制和管理，确保不同用户或系统只能访问和操作其被授权的部分，从而维护数据的安全性和合规性。\"\n"
            "            }\n"
            "          ]\n"
            "        },\n"
            "        {\n"
            "          \"function_name\": \"角色管理\",\n"
            "          \"function_pages\": [\n"
            "            {\n"
            "              \"page_name\": \"角色管理页面\",\n"
            "              \"page_overview\": \"角色管理页面的主要功能在于对系统中的角色进行维护、授权和管理，以确保用户只能访问他们所需的功"
            "            }\n"
            "          ]\n"
            "        }\n"
            "      ]\n"
            "    }\n"
            "  ]\n"
            "}"
        )

        print("开始生成功能清单...")
        # 使用LLM生成JSON
        func_names_json_response = self.llm.predict(func_names_json_prompt)
        
        try:
            # Strip leading/trailing whitespace and newlines
            stripped_response = func_names_json_response.strip().strip('```json')

            # Parse the stripped response into JSON
            func_names_info = json.loads(stripped_response)

            # Use func_names_info as needed
            print("功能详情:", func_names_info)

        except json.JSONDecodeError as e:
            # Handle JSON decoding errors
            print(f"JSON decoding error: {e}")
            func_names_info = None  # Or set to a default value
            logger.error("Traceback details:", exc_info=True)

        except Exception as e:
            # Handle any other unexpected errors
            print(f"An error occurred: {e}")
            func_names_info = None  # Or set to a default value
            logger.error("Traceback details:", exc_info=True)
        try:
            self.save_func_names_json_to_excel(func_names_info, "./功能清单.xlsx")
            print(f"功能清单已保存")
        except json.JSONDecodeError as e:
            print(f"JSON 解码错误: {e}")
            func_names_json = None

if __name__ == "__main__":
    # Define the list of platforms

    # Path to the product overview
    overview_path = "./overview.txt"

    generator = DataPreprocessingModuleSpecGenerator(
            overview_path=overview_path,
        )
    try:
        generator.process()
        logger.info(f"成功生成功能清单")
    except Exception as e:
        logger.error(f"生成功能清单失败！")
        logger.error("Traceback details:", exc_info=True)